import os
from openpyxl import load_workbook

# 目标文件夹路径
base_folder = r'D:\All_of_mine\大学\项目和比赛\大创\src\data'

# 遍历目标文件夹中的每个子文件夹
for folder_name in os.listdir(base_folder):
    folder_path = os.path.join(base_folder, folder_name)

    # 检查是否是文件夹
    if os.path.isdir(folder_path):
        # 检查文件夹内是否存在 stock_info.xlsx 文件
        file_path = os.path.join(folder_path, 'stock_info.xlsx')

        if os.path.exists(file_path):
            # 打开 Excel 文件并读取 URL 列
            wb_stock = load_workbook(file_path)
            ws_stock = wb_stock.active

            # 获取“URL”列的索引，假设 URL 列在第一行的“URL”列名下
            header = [cell.value for cell in ws_stock[1]]  # 获取表头
            if "URL" in header:
                url_index = header.index("URL")  # 获取 URL 列的索引
            else:
                continue  # 如果没有找到“URL”列，跳过该文件

            # 定义输出文件路径
            guba_output_file = os.path.join(folder_path, 'guba_urls.txt')
            caifu_output_file = os.path.join(folder_path, 'caifu_urls.txt')

            # 打开两个文件，用于存储筛选后的 URL
            with open(guba_output_file, 'w', encoding='utf-8') as guba_file, open(caifu_output_file, 'w',
                                                                                  encoding='utf-8') as caifu_file:

                # 筛选并写入包含“guba”或“caifu”的 URL
                for row in ws_stock.iter_rows(min_row=2, values_only=True):
                    url = row[url_index]
                    if url:
                        if "guba" in url:
                            guba_file.write(url + '\n')  # 将包含 guba 的 URL 写入 guba.txt
                        elif "caifu" in url:
                            caifu_file.write(url + '\n')  # 将包含 caifu 的 URL 写入 caifu.txt

            print(f"股票代码 {folder_name} 的 URL 筛选完成，结果已保存到 {folder_path}")
