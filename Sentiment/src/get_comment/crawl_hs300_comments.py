import requests
from lxml import etree
from openpyxl import Workbook
import os
import time
from datetime import datetime

from openpyxl.reader.excel import load_workbook

from src.cai_fu_comments.code.check_time import check_post_time_and_continue

# 设置 HTTP 请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': 'https://guba.eastmoney.com/',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'close'  # 禁用 keep-alive 以便代理切换 IP
}

# 快代理隧道信息
TUNNEL = ""
USERNAME = "t14514172566719"
PASSWORD = ""

# 配置代理
proxies = {
    "http": f"http://{USERNAME}:{PASSWORD}@{TUNNEL}/",
    "https": f"http://{USERNAME}:{PASSWORD}@{TUNNEL}/"
}

# 沪深300股票代码（简化示例，按需填写完整）
stock_codes = ["zssh000300"]

# 获取网页内容
def fetch_page(url):
    try:
        response = requests.get(url, headers=headers, proxies=proxies, timeout=10)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        print(f"请求失败: {e}，10 秒后重试...")
        return None

# 从 URL 页面提取时间
def fetch_time_from_url(url):
    """ 访问 URL 并提取时间信息 """
    html = fetch_page(url)
    if not html:
        return None

    tree = etree.HTML(html)
    date_str = tree.xpath('//div[@class="publish_time"]/text()')  # 解析时间
    print(date_str)
    if date_str:
        date_str = date_str[0].strip()
        try:
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")  # 转换为 datetime 格式
        except ValueError:
            print(f"日期格式错误: {date_str}")
            return None
    return None

# 解析 HTML 并提取数据
def parse_html(html):
    tree = etree.HTML(html)
    items = tree.xpath('//tbody[@class="listbody"]/tr')

    data_list = []
    for item in items:
        read = item.xpath('.//div[@class="read"]/text()')
        read = read[0].strip() if read else 'error'

        comment = item.xpath('.//div[@class="reply"]/text()')
        comment = comment[0].strip() if comment else 'error'

        title = item.xpath('.//div[@class="title"]//a/text()')
        title = title[0].strip() if title else ' error'

        author = item.xpath('.//div[@class="author"]/a/text()')
        author = author[0].strip() if author else 'error'

        update = item.xpath('.//div[@class="update"]/text()')
        update = update[0].strip() if update else 'error'

        link = item.xpath('.//div[@class="title"]//a/@href')
        if link:
            link = link[0].strip()
            if link.startswith('/news'):
                link = "https://guba.eastmoney.com" + link
            elif link.startswith('//caifuhao.eastmoney.com'):
                link = "https:" + link
            else:
                link = 'error'
        else:
            link = 'error'

        data_list.append([title, read, comment, author, update, link])
    return data_list

# 主爬取流程
def main():
    start_page = 400
    end_page = 1000
    start_time = time.time()

    # 使用沪深300的股票代码
    for stock_code in stock_codes:
        print(f"正在爬取股票代码 {stock_code} 的数据...")
        s_time = time.time()

        folder_name = os.path.join(r"D:\All_of_mine\大学\项目和比赛\da_chuang\src\data_2", stock_code)
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        file_name = os.path.join(folder_name, 'stock_info.xlsx')
        file_name = os.path.join(folder_name, 'stock_info.xlsx')

        # 如果文件存在，加载；否则新建并加表头
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Info"
            ws.append(["标题", "阅读量", "评论", "作者", "最后更新时间", "URL"])

        stop_scraping = False  # 标记是否提前停止

        for page in range(start_page, end_page + 1):
            url = f'https://guba.eastmoney.com/list,{stock_code}_{page}.html?returnCode=0'
            print(f"正在爬取第 {page} 页...")
            html = fetch_page(url)

            if html:
                data_list = parse_html(html)
                for row in data_list:
                    ws.append(row)

                print(f"股票代码 {stock_code} 第 {page} 页爬取成功！")

                if page % 40 == 0:
                    print("检查 URL 时间...")
                    for row in data_list:
                        url_to_check = row[-1]  # 最后一列是 URL
                        if url_to_check != 'error':
                            is_continue = check_post_time_and_continue(url_to_check)
                            if not is_continue:
                                print(f"跳过 {stock_code}")
                                stop_scraping = True
                                break
                            print(f"继续 {stock_code}")

                if stop_scraping:
                    break  # 直接跳过剩余页面，爬取下一个股票
            else:
                print(f"股票代码 {stock_code} 第 {page} 页爬取失败，10 秒后重试！")

        wb.save(file_name)

        e_time = time.time()
        print(f"股票代码 {stock_code} 数据已保存。用时{e_time - s_time:.2f}秒\n")

    end_time = time.time()
    print(f"所有数据爬取完成！总运行时间：{(end_time - start_time) / 60:.2f} 分钟")

if __name__ == "__main__":
    main()
