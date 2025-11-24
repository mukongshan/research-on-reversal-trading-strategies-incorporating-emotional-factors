import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
import os
import time
import random
from datetime import datetime

from src.cai_fu_comments.code.check_time import check_post_time_and_continue

# 设置 HTTP 请求头，模拟浏览器访问
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': 'https://guba.eastmoney.com/',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'close'  # 禁用 keep-alive 以便代理切换 IP
}

# 快代理隧道信息
TUNNEL = "c920.kdltps.com:15818"
USERNAME = "t14152351753831"
PASSWORD = "nhenjhtk"

# 配置代理
proxies = {
    "http": f"http://{USERNAME}:{PASSWORD}@{TUNNEL}/",
    "https": f"http://{USERNAME}:{PASSWORD}@{TUNNEL}/"
}

# 读取 Excel 文件中的股票代码
def read_stock_codes(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    header = [cell.value for cell in sheet[1]]
    stock_code_index = header.index("成份券代码Constituent Code")
    stock_codes = [str(row[stock_code_index]) for row in sheet.iter_rows(min_row=2, max_col=stock_code_index + 1, values_only=True)]
    return stock_codes

# 获取网页内容
def fetch_page(url):
    try:
        response = requests.get(url, headers=headers, proxies=proxies, timeout=10)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        print(f"请求失败: {e}，10 秒后重试...")
        return None

# 从 URL 页面提取时间
def fetch_time_from_url(url):
    """ 访问 URL 并提取时间信息 """
    html = fetch_page(url)
    if not html:
        return None

    tree = etree.HTML(html)
    date_str = tree.xpath('//div[@class="publish_time"]/text()')  # 解析时间
    print(date_str)
    if date_str:
        date_str = date_str[0].strip()
        try:
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")  # 转换为 datetime 格式
        except ValueError:
            print(f"日期格式错误: {date_str}")
            return None
    return None

# 解析 HTML 并提取数据
def parse_html(html):
    tree = etree.HTML(html)
    items = tree.xpath('//tbody[@class="listbody"]/tr')

    data_list = []
    for item in items:
        read = item.xpath('.//div[@class="read"]/text()')
        read = read[0].strip() if read else 'error'

        comment = item.xpath('.//div[@class="reply"]/text()')
        comment = comment[0].strip() if comment else 'error'

        title = item.xpath('.//div[@class="title"]//a/text()')
        title = title[0].strip() if title else ' error'

        author = item.xpath('.//div[@class="author"]/a/text()')
        author = author[0].strip() if author else 'error'

        update = item.xpath('.//div[@class="update"]/text()')
        update = update[0].strip() if update else 'error'

        link = item.xpath('.//div[@class="title"]//a/@href')
        if link:
            link = link[0].strip()
            if link.startswith('/news'):
                link = "https://guba.eastmoney.com" + link
            elif link.startswith('//caifuhao.eastmoney.com'):
                link = "https:" + link
            else:
                link = 'error'
        else:
            link = 'error'

        data_list.append([title, read, comment, author, update, link])
    return data_list

# 主爬取流程
def main():
    start_page = 1
    end_page = 1000000
    start_time = time.time()

    file_path = r"D:\All_of_mine\大学\项目和比赛\da_chuang\src\data\股票代码.xlsx"
    wb_stock = load_workbook(file_path)
    sheet_stock = wb_stock.active

    # 读取表头，找到需要的列索引
    header = [cell.value for cell in sheet_stock[1]]
    stock_code_index = header.index("成份券代码Constituent Code")
    read_status_index = header.index("是否读取")  # 确保表格里有这一列

    # 读取所有股票代码及其读取状态
    stock_data = [(row[stock_code_index], row[read_status_index]) for row in sheet_stock.iter_rows(min_row=2, values_only=True)]

    for row_idx, (stock_code, read_status) in enumerate(stock_data, start=2):  # 从 Excel 第二行开始
        if str(read_status).strip() == "是":
            print(f"股票代码 {stock_code} 已读取，跳过...")
            continue  # 跳过已爬取的股票

        print(f"正在爬取股票代码 {stock_code} 的数据...")
        s_time = time.time()

        folder_name = os.path.join(r"D:\All_of_mine\大学\项目和比赛\da_chuang\src\data", stock_code)
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        file_name = os.path.join(folder_name, 'stock_info.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Info"
        ws.append(["标题", "阅读量", "评论", "作者", "最后更新时间", "URL"])

        stop_scraping = False  # 标记是否提前停止

        for page in range(start_page, end_page + 1):
            url = f'https://guba.eastmoney.com/list,{stock_code}_{page}.html?returnCode=0'
            html = fetch_page(url)

            if html:
                data_list = parse_html(html)
                for row in data_list:
                    ws.append(row)

                print(f"股票代码 {stock_code} 第 {page} 页爬取成功！")

                if page % 40 == 0:
                    print("检查 URL 时间...")
                    for row in data_list:
                        url_to_check = row[-1]  # 最后一列是 URL
                        if url_to_check != 'error':
                            is_continue = check_post_time_and_continue(url_to_check)
                            if not is_continue:
                                print(f"跳过 {stock_code}")
                                stop_scraping = True
                                break
                            print(f"继续 {stock_code}")

                if stop_scraping:
                    break  # 直接跳过剩余页面，爬取下一个股票
            else:
                print(f"股票代码 {stock_code} 第 {page} 页爬取失败，10 秒后重试！")

        wb.save(file_name)

        # 在 Excel 中标记已读取
        sheet_stock.cell(row=row_idx, column=read_status_index + 1, value="是")
        wb_stock.save(file_path)

        e_time = time.time()
        print(f"股票代码 {stock_code} 数据已保存。用时{e_time - s_time:.2f}秒\n")

    end_time = time.time()
    print(f"所有数据爬取完成！总运行时间：{(end_time - start_time) / 60:.2f} 分钟")

if __name__ == "__main__":
    main()
